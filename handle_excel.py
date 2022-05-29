import os
import logging
import openpyxl
from openpyxl import Workbook


class ExcelHandle(object):

    def __init__(self, name, upload_file):
        self.name = name
        self.upload_file = upload_file
        self.result_excel_header = ['导入编号', '网店订单号', '下单时间', '付款时间', '承诺发货时间', '客户账号', '客户名称',
                                    '客户邮箱', 'QQ', '收货人', '手机', '固定电话', '国家', '省份', '市（区）', '区（县）',
                                    '收货地址', '邮政编码', '发货仓库', '应收邮资', '平台佣金', '客付税额', '应收合计', '客服备注',
                                    '客户备注', '销售渠道名称', '结算方式', '货品名称', '条码', '货品编号', '规格', '批次号',
                                    '数量', '单价', '货品优惠', '金额', '网店子订单号', '定制码', '货品备注', '发票抬头',
                                    '发票类型', '证件类型', '证件号码', '证件使用姓名', '物流公司', '物流单号', '支付单号',
                                    '收款账户', '业务员', '跟单员', '标记']

    def init_read_excel(self):
        """
        初始化而excel并插入表头数据
        :return:
        """
        workbook = openpyxl.load_workbook(self.upload_file)
        # 获取第一个 sheet 表格
        sheet_name = workbook[workbook.sheetnames[0]]
        wb = Workbook()
        ws = wb.active
        ws.append(self.result_excel_header)
        return wb, ws, list(sheet_name.rows)[1:]

    def delete_success_file(self):
        """
        处理完之后删除excel
        :return:
        """
        if os.path.exists(self.upload_file):
            os.remove(self.upload_file)
            logging.info("delete upload file {0} success".format(self.upload_file))
        else:
            logging.warning("not found upload file {0}", self.upload_file)

