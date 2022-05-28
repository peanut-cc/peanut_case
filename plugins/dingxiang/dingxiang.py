import logging
import os
import openpyxl
from openpyxl import Workbook
from handle_excel import ExcelHandle
import time


class CustomExcelHandle(ExcelHandle):

    def __init__(self, name, upload_file):
        super().__init__(name, upload_file)
        self.result_excel_header = ["导入编号", "网店订单号", "客户账号", "客户名销售渠道名称称", "收货人", "手机", "省份",
                                    "市（区）", "区（县）", "收货地址", "销售渠道名称", "货品名称", "条码", "数量", "单价"]
        self.custom_name = "丁香（无痕）"

    def start_handle_excel(self):
        print("{} start handle excel".format(self.name))
        workbook = openpyxl.load_workbook(self.upload_file)
        # 获取第一个 sheet 表格
        sheet_name = workbook[workbook.sheetnames[0]]
        wb = Workbook()
        ws = wb.active
        ws.append(self.result_excel_header)
        for index, row in enumerate(list(sheet_name.rows)[1:]):
            sn = row[2].value
            shop_sn = row[2].value
            receive_people = row[19].value
            phone = row[20].value
            province = row[21].value
            city = row[22].value
            county = row[23].value
            address = row[24].value
            sales_channel_name = self.custom_name
            product_name = row[10].value
            barcode = row[13].value
            numbers = row[15].value
            unit_price = row[18].value
            ws.append([sn, shop_sn, "", self.custom_name, receive_people, phone, province, city, county, address,
                       sales_channel_name, product_name, barcode, numbers, unit_price])

        result_name = "./result/dingxiang/丁香{0}.xlsx".format(time.strftime("%Y%m%d%H%M%S", time.localtime()))
        wb.save(result_name)
        logging.info("{0} handle excel file {1} success".format(self.name, self.upload_file))
        self.delete_success_file()

    def delete_success_file(self):
        if os.path.exists(self.upload_file):
            os.remove(self.upload_file)
            logging.info("delete upload file {0} success".format(self.upload_file))
        else:
            logging.warning("not found upload file {0}", self.upload_file)











